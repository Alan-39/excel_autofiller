import os
from openpyxl import load_workbook

def read_data(row_idx, col_idx):
    data_files = os.listdir("data")
    data_files.sort()

    data_list = []
    data_entry = 0
    print(data_files)
    for i in range(len(data_files)):
        wb = load_workbook("data/{}".format(data_files[i]))
        sheet = wb.active
        row = row_idx
        while True:
            if sheet.cell(row=row, column=col_idx).value == None:
                break
            else:
                data_list.append(sheet.cell(row, column=col_idx).value)
                row = row + 1
                data_entry = data_entry + 1
    print("{} Entries found".format(data_entry))
    return data_list

def write_data(target_sheet_list, target_list, data_list):
    wb = load_workbook("target/{}".format(os.listdir("target")[0]))
    
    for sheet in range(len(target_sheet_list)):
        wb_sheet = wb[target_sheet_list[sheet]]
        curr_idx = 0
        parity_list = data_list[sheet::2]

        for n in range(len(target_list)):
            for i in range(target_list[n][1], target_list[n][2] + 1):
                wb_sheet.cell(row=i, column=target_list[n][0]).value = parity_list[curr_idx]
                curr_idx = curr_idx + 1
    wb.save("target/{}".format(os.listdir("target")[0]))

def main():
    target_sheet_list = []

    print(
        "====== Excel Autofiller by Alan =====\n"
        "This program takes a list of data from csv file and populates it to the target excel\n\n"
        "Put the data entries u wish to export on folder named data, and entries u wish to import on folder name target"
        )
    while True:
        try:
            data_row = int(input("Enter starting row: "))
            data_col = int(input("Enter starting column: "))
            break
        except ValueError:
            print("All specified row & column MUST be in numerials>>!!\n")
    horz_sheet = str(input("Enter horz sheetname: "))
    vert_sheet = str(input("Enter vert sheetname: "))
    target_sheet_list.append(vert_sheet)
    target_sheet_list.append(horz_sheet)

    data_list = read_data(data_row, data_col)
    target_list = [] # [target col, start row, end row]
    while True:
        try:
            target_col = int(input("Enter target column: "))
            start_row = int(input("Enter starting row: "))
            end_row = int(input("Enter ending row: "))

            target_list.append([target_col, start_row, end_row])
            prompt = input("Add more range? (y/n): ").lower()
            if prompt == 'n':
                break
        except ValueError:
            print("All specified row & column MUST be in numerials>>!!\n")
    write_data(target_sheet_list, target_list, data_list)
    pexit = input("Autofilled! enter any key to exit")

if __name__ == "__main__":
    main()